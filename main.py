from flask import Flask, render_template, request
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import xlrd
import os
import unicodedata
import re

from config import Config
from models import db, ImportHistory, VehicleExpense

app = Flask(__name__)
app.config.from_object(Config)
def format_ron(value):
    if value is None:
        value = 0

    formatted = f"{value:,.2f}"
    formatted = formatted.replace(",", "X")
    formatted = formatted.replace(".", ",")
    formatted = formatted.replace("X", ".")

    return formatted


app.jinja_env.filters["ron"] = format_ron

db.init_app(app)

ALLOWED_EXTENSIONS = {"xls", "xlsx"}

REQUIRED_COLUMNS = [
    "numar",
    "marca",
    "model",
    "locatie",
    "sofer",
    "revizii",
    "reparatii",
    "carburant",
    "anvelope",
]

COLUMN_ALIASES = {
    "numar": ["numar", "număr", "numar de inmatriculare", "număr de înmatriculare"],
    "marca": ["marca", "marcă"],
    "model": ["model"],
    "tip": ["tip"],
    "departament": ["departament"],
    "locatie": ["locatie", "locație"],
    "centru_cost": ["centru de cost", "centru cost"],
    "entitate": ["entitate"],
    "status": ["status"],
    "sofer": ["sofer", "șofer"],
    "casco": ["casco"],
    "rca": ["rca"],
    "impozite": ["impozite"],
    "roviniete": ["roviniete"],
    "itp": ["itp"],

    "revizii": ["revizii"],
    "reparatii": ["reparatii", "reparații"],
    "carburant": ["carburant"],
    "anvelope": ["anvelope"],
    "acumulatori": ["acumulatori"],
    "accident": ["accident"],
    "amenzi": ["amenzi"],
    "alte_cheltuieli": ["alte cheltuieli", "alte chelt"],
    "retineri": ["retineri", "rețineri"],
    "rate": ["rate"],
    "amortizari": ["amortizari", "amortizări"],
}


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def normalize_text(value):
    if value is None:
        return ""

    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return text.strip()


def safe_float(value):
    if value is None or str(value).strip() == "":
        return 0

    try:
        text = str(value).strip()
        text = text.replace(" ", "")

        if "," in text:
            text = text.replace(".", "")
            text = text.replace(",", ".")

        return float(text)
    except Exception:
        return 0


def read_excel_rows(file_path):
    extension = file_path.rsplit(".", 1)[1].lower()

    if extension == "xlsx":
        workbook = load_workbook(file_path, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        return worksheet.title, rows

    if extension == "xls":
        workbook = xlrd.open_workbook(file_path)
        worksheet = workbook.sheet_by_index(0)

        rows = []
        for row_idx in range(worksheet.nrows):
            rows.append(worksheet.row_values(row_idx))

        return worksheet.name, rows

    return None, []


def find_header_row(rows, max_scan_rows=30):
    best_index = None
    best_match_count = 0
    best_headers = []

    for i in range(min(len(rows), max_scan_rows)):
        current_row = [normalize_text(cell) for cell in rows[i]]

        match_count = 0

        for aliases in COLUMN_ALIASES.values():
            normalized_aliases = [normalize_text(alias) for alias in aliases]

            if any(alias in current_row for alias in normalized_aliases):
                match_count += 1

        if match_count > best_match_count:
            best_match_count = match_count
            best_index = i
            best_headers = rows[i]

    return best_index, best_headers, best_match_count


def build_column_map(headers):
    column_map = {}
    normalized_headers = [normalize_text(header) for header in headers]

    for canonical_name, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            normalized_alias = normalize_text(alias)

            if normalized_alias in normalized_headers:
                column_map[canonical_name] = normalized_headers.index(normalized_alias)
                break

    return column_map


def get_cell(row, column_map, field_name):
    index = column_map.get(field_name)

    if index is None or index >= len(row):
        return None

    value = row[index]
    return None if value == "" else value


@app.route("/")
def home():
    records = VehicleExpense.query.all()

    total_inregistrari = len(records)

    total_reparatii = sum(row.total_reparatii or 0 for row in records)
    total_carburant = sum(row.carburant or 0 for row in records)
    total_taxe = sum(row.total_taxe or 0 for row in records)
    total_anvelope = sum(row.anvelope or 0 for row in records)
    total_general = sum(row.total_general or 0 for row in records)

    chart_labels = ["Reparații", "Carburant", "Taxe", "Anvelope"]
    chart_values = [
        round(total_reparatii, 2),
        round(total_carburant, 2),
        round(total_taxe, 2),
        round(total_anvelope, 2),
    ]

    return render_template(
        "index.html",
        total_inregistrari=total_inregistrari,
        total_reparatii=round(total_reparatii, 2),
        total_carburant=round(total_carburant, 2),
        total_taxe=round(total_taxe, 2),
        total_anvelope=round(total_anvelope, 2),
        total_general=round(total_general, 2),
        chart_labels=chart_labels,
        chart_values=chart_values,
    )


@app.route("/import", methods=["GET", "POST"])
def import_excel():
    message = None
    message_type = None
    preview_headers = []
    preview_rows = []
    uploaded_filename = None
    sheet_name = None
    missing_columns = []
    detected_columns = []
    total_data_rows = 0
    header_row_index = None

    recent_imports = ImportHistory.query.order_by(ImportHistory.id.desc()).limit(10).all()

    if request.method == "POST":
        if "excel_file" not in request.files:
            message = "Nu a fost trimis niciun fișier."
            message_type = "error"

        else:
            file = request.files["excel_file"]

            if file.filename == "":
                message = "Te rog selectează un fișier."
                message_type = "error"

            elif not allowed_file(file.filename):
                message = "Sunt acceptate doar fișiere .xls și .xlsx"
                message_type = "error"

            else:
                filename = secure_filename(file.filename)
                save_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
                file.save(save_path)

                try:
                    sheet_name, rows = read_excel_rows(save_path)
                    uploaded_filename = filename

                    if not rows:
                        message = "Fișierul Excel este gol."
                        message_type = "error"

                    else:
                        header_row_index, detected_headers, match_count = find_header_row(rows)

                        if header_row_index is None or match_count == 0:
                            message = "Nu am putut identifica automat header-ul din fișier."
                            message_type = "error"

                        else:
                            preview_headers = [
                                str(cell).strip() if cell is not None else ""
                                for cell in detected_headers
                            ]

                            column_map = build_column_map(preview_headers)
                            detected_columns = list(column_map.keys())

                            missing_columns = [
                                col for col in REQUIRED_COLUMNS if col not in column_map
                            ]

                            data_rows = rows[header_row_index + 1:]

                            non_empty_data_rows = [
                                row for row in data_rows
                                if any(
                                    cell is not None and str(cell).strip() != ""
                                    for cell in row
                                )
                            ]

                            # Filtreaza randurile fara numar de inmatriculare valid
                            def is_valid_plate(value):
                                if not value:
                                    return False
                                text = str(value).strip()
                                return bool(re.match(r'^[A-Z]{1,3}\d{2,3}[A-Z]{2,3}$', text, re.IGNORECASE))

                            numar_idx = column_map.get("numar")
                            if numar_idx is not None:
                                non_empty_data_rows = [
                                    row for row in non_empty_data_rows
                                    if numar_idx < len(row) and is_valid_plate(row[numar_idx])
                                ]

                            total_data_rows = len(non_empty_data_rows)

                            max_preview_rows = 20
                            for row in non_empty_data_rows[:max_preview_rows]:
                                preview_rows.append([
                                    "" if cell is None else str(cell)
                                    for cell in row
                                ])

                            if missing_columns:
                                status = "INVALID"
                                message = "Header-ul a fost identificat, dar lipsesc unele coloane obligatorii."
                                message_type = "error"

                            else:
                                status = "SUCCESS"
                                message = f'Fișierul "{filename}" a fost încărcat, validat și salvat cu succes.'
                                message_type = "success"

                            import_record = ImportHistory(
                                filename=filename,
                                sheet_name=sheet_name,
                                total_rows=total_data_rows,
                                status=status
                            )

                            db.session.add(import_record)
                            db.session.commit()
                            
                            VehicleExpense.query.delete()
                            db.session.commit()

                            if not missing_columns:
                                for row in non_empty_data_rows:
                                    revizii = safe_float(get_cell(row, column_map, "revizii"))
                                    reparatii = safe_float(get_cell(row, column_map, "reparatii"))
                                    carburant = safe_float(get_cell(row, column_map, "carburant"))
                                    anvelope = safe_float(get_cell(row, column_map, "anvelope"))
                                    acumulatori = safe_float(get_cell(row, column_map, "acumulatori"))
                                    accident = safe_float(get_cell(row, column_map, "accident"))
                                    amenzi = safe_float(get_cell(row, column_map, "amenzi"))
                                    alte_cheltuieli = safe_float(get_cell(row, column_map, "alte_cheltuieli"))
                                    casco = safe_float(get_cell(row, column_map, "casco"))
                                    rca = safe_float(get_cell(row, column_map, "rca"))
                                    impozite = safe_float(get_cell(row, column_map, "impozite"))
                                    roviniete = safe_float(get_cell(row, column_map, "roviniete"))
                                    itp = safe_float(get_cell(row, column_map, "itp"))
                                    retineri = safe_float(get_cell(row, column_map, "retineri"))
                                    rate = safe_float(get_cell(row, column_map, "rate"))
                                    amortizari = safe_float(get_cell(row, column_map, "amortizari"))

                                    total_reparatii = revizii + reparatii + acumulatori + alte_cheltuieli
                                    total_taxe = casco + rca + impozite + roviniete
                                    total_general = total_reparatii + carburant + total_taxe + anvelope

                                    vehicle_expense = VehicleExpense(
                                        import_id=import_record.id,

                                        numar=str(get_cell(row, column_map, "numar") or ""),
                                        marca=str(get_cell(row, column_map, "marca") or ""),
                                        model=str(get_cell(row, column_map, "model") or ""),
                                        tip=str(get_cell(row, column_map, "tip") or ""),
                                        departament=str(get_cell(row, column_map, "departament") or ""),
                                        locatie=str(get_cell(row, column_map, "locatie") or ""),
                                        centru_cost=str(get_cell(row, column_map, "centru_cost") or ""),
                                        entitate=str(get_cell(row, column_map, "entitate") or ""),
                                        status=str(get_cell(row, column_map, "status") or ""),
                                        sofer=str(get_cell(row, column_map, "sofer") or ""),

                                        revizii=revizii,
                                        reparatii=reparatii,
                                        carburant=carburant,
                                        anvelope=anvelope,
                                        acumulatori=acumulatori,
                                        accident=accident,
                                        amenzi=amenzi,
                                        alte_cheltuieli=alte_cheltuieli,
                                        retineri=retineri,
                                        rate=rate,
                                        amortizari=amortizari,
                                        casco=casco,
                                        rca=rca,
                                        impozite=impozite,
                                        roviniete=roviniete,
                                        itp=itp,

                                        total_reparatii=total_reparatii,
                                        total_taxe=total_taxe,
                                        total_general=total_general,
                                    )

                                    db.session.add(vehicle_expense)

                                db.session.commit()

                except Exception as e:
                    message = f"A apărut o eroare la citirea fișierului: {str(e)}"
                    message_type = "error"

    recent_imports = ImportHistory.query.order_by(ImportHistory.id.desc()).limit(10).all()

    return render_template(
        "import_excel.html",
        message=message,
        message_type=message_type,
        preview_headers=preview_headers,
        preview_rows=preview_rows,
        uploaded_filename=uploaded_filename,
        sheet_name=sheet_name,
        missing_columns=missing_columns,
        detected_columns=detected_columns,
        required_columns=REQUIRED_COLUMNS,
        total_data_rows=total_data_rows,
        header_row_index=header_row_index,
        recent_imports=recent_imports,
    )

@app.route("/monitorizare")
def monitorizare():
    locatie_filter = request.args.get("locatie", "").strip()
    sofer_filter = request.args.get("sofer", "").strip()
    numar_filter = request.args.get("numar", "").strip()
    centru_cost_filter = request.args.get("centru_cost", "").strip()

    query = VehicleExpense.query

    if locatie_filter:
        query = query.filter(VehicleExpense.locatie.ilike(f"%{locatie_filter}%"))

    if sofer_filter:
        query = query.filter(VehicleExpense.sofer.ilike(f"%{sofer_filter}%"))

    if numar_filter:
        query = query.filter(VehicleExpense.numar.ilike(f"%{numar_filter}%"))

    if centru_cost_filter:
        query = query.filter(VehicleExpense.centru_cost.ilike(f"%{centru_cost_filter}%"))

    records = query.order_by(VehicleExpense.id.desc()).all()

    total_reparatii = sum(row.total_reparatii or 0 for row in records)
    total_carburant = sum(row.carburant or 0 for row in records)
    total_taxe = sum(row.total_taxe or 0 for row in records)
    total_anvelope = sum(row.anvelope or 0 for row in records)
    total_general = sum(row.total_general or 0 for row in records)

    all_records = VehicleExpense.query.all()

    numar_options = sorted({row.numar for row in all_records if row.numar})
    locatie_options = sorted({row.locatie for row in all_records if row.locatie})
    sofer_options = sorted({row.sofer for row in all_records if row.sofer})
    centru_cost_options = sorted({row.centru_cost for row in all_records if row.centru_cost})

    return render_template(
        "monitorizare.html",
        records=records,
        locatie_filter=locatie_filter,
        sofer_filter=sofer_filter,
        numar_filter=numar_filter,
        centru_cost_filter=centru_cost_filter,
        total_reparatii=round(total_reparatii, 2),
        total_carburant=round(total_carburant, 2),
        total_taxe=round(total_taxe, 2),
        total_anvelope=round(total_anvelope, 2),
        total_general=round(total_general, 2),
        numar_options=numar_options,
        locatie_options=locatie_options,
        sofer_options=sofer_options,
        centru_cost_options=centru_cost_options,
    )


if __name__ == "__main__":
    with app.app_context():
        db.create_all()

    app.run(debug=True)

    